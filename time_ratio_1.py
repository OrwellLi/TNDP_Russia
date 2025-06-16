from openpyxl import load_workbook
import numpy as np
import copy as copy
import pandas as pd


wb_Airport_data = load_workbook(r'TNDP_Russia/Support_data/Airport_data.xlsx')
ws_Airport_data = wb_Airport_data['Sheet1']
Airport_data = np.array([[i.value for i in j] for j in ws_Airport_data['B2':'H14174']])  

set_of_airports = set() 

for row in range(len(Airport_data)-1):
    # Добавляем только не-None значения
    if Airport_data[row+1][2] is not None and str(Airport_data[row+1][2]).strip() != '':
        set_of_airports.add(str(Airport_data[row+1][2]).strip())
    if Airport_data[row+1][3] is not None and str(Airport_data[row+1][3]).strip() != '':
        set_of_airports.add(str(Airport_data[row+1][3]).strip())

set_of_airports.discard('None')
set_of_airports.discard('')
# print(set_of_airports)
    
Num_airports  =  len(set_of_airports)

list_Airport_temp = sorted(list(set_of_airports))

list_of_airports = np.zeros(Num_airports, dtype=object)

for n in range(Num_airports):
    list_of_airports[n] = list_Airport_temp[n]


wb_Aiport_info = load_workbook(r'TNDP_Russia/Source_data/airportinformation3.xlsx')
ws_Airport_info = wb_Aiport_info['airportdata']

name = np.array([[i.value for i in j] for j in ws_Airport_info['B2':'GPG2']]) 
XX = np.zeros(len(name[0]), dtype=object)
for i in range(len(XX)):
    XX[i] = str(copy.copy(name[0,i]))
name = XX #Название
# print("")
# print(name)

country = np.array([[i.value for i in j] for j in ws_Airport_info['B4':'GPG4']]) 
XX = np.zeros(len(country[0]), dtype=object)
for i in range(len(XX)):
    XX[i] = copy.copy(country[0,i])
country = XX #Страна
# print("")
# print(country)

IATA = np.array([[i.value for i in j] for j in ws_Airport_info['B7':'GPG7']]) 
XX = np.zeros(len(IATA[0]), dtype=object)
for i in range(len(XX)):
    XX[i] = str(copy.copy(IATA[0,i]))
IATA = XX #ИАТА-код
# print("")
# print(IATA)

Airport_longitude = np.array([[i.value for i in j] for j in ws_Airport_info['B8':'GPG8']]) 
XX = np.zeros(len(Airport_longitude[0]))
for i in range(len(XX)):
    XX[i] = str(copy.copy(Airport_longitude[0,i]))
Airport_longitude = XX #широта
# print("")
# print(Airport_longitude)

Airport_latitude = np.array([[i.value for i in j] for j in ws_Airport_info['B9':'GPG9']]) 
XX = np.zeros(len(Airport_latitude[0]))
for i in range(len(XX)):
    XX[i] = str(copy.copy(Airport_latitude[0,i]))
Airport_latitude = XX #долгота
# print("")
# print(Airport_latitude)

airport_information = np.full((Num_airports, 5),'n/a', dtype=object) #[IATA, name, coun, AP_lat, AP_lon]

#заполняем таблицу
for row in range(len(list_of_airports)):
    airport_information[row][0] = list_of_airports[row]
    for column in range(len(name)):
        if IATA[column] == airport_information[row][0]:
            # airport_information[row][1] = IATA[column]
            airport_information[row][1] = name[column]
            airport_information[row][2] = country[column]
            airport_information[row][3] = Airport_latitude[column]
            airport_information[row][4] = Airport_longitude[column]
            

#обновляем последовательность ап в коде IATA
for n in range(Num_airports):
    list_of_airports[n] = airport_information[n][0]


# Пути к файлам Excel
file_path_1 = "TNDP_Russia/Total_output_data/DEMAND_TOTAL_TRIP_TIME_standart.xlsx"  # Первый файл с временем пути
file_path_2 = "TNDP_Russia/Total_output_data/DEMAND_TOTAL_TRIP_TIME.xlsx"  # Второй файл с временем пути
output_file = "TNDP_Russia/Total_output_data/TIME_RATIO.xlsx"  # Выходной файл с отношением

# Чтение файлов в DataFrame
df1 = pd.read_excel(file_path_1, index_col=0)  # Устанавливаем первый столбец как индекс (аэропорты)
df2 = pd.read_excel(file_path_2, index_col=0)  # Устанавливаем первый столбец как индекс (аэропорты)

# Проверка, что таблицы имеют одинаковые размеры и индексы
if not df1.index.equals(df2.index) or not df1.columns.equals(df2.columns):
    raise ValueError("log. Таблицы имеют разные индексы или столбцы. Убедитесь, что аэропорты совпадают.")

# Вычисление отношения df1 / df2, заменяя деление на ноль и NaN на 0 или другое значение по умолчанию
ratio_matrix = df2.div(df1, fill_value=0).replace([np.inf, -np.inf], 0).fillna(0)

# Сохранение результата в новый Excel файл
ratio_matrix.to_excel(output_file)

print()
print(f"log. Отношение первой таблицы ко второй сохранено в {output_file}")
print()