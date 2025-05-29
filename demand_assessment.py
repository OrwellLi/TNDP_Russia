#aopa

"""
Created on Thu 22 MAY 16:22:22 2025

@author OrwellLi

Project based on solving Transit Network Design Problem (TNDP) and then will be optimized

"""
#---------------------------------
#Импорт библиотек и настроек

from openpyxl import load_workbook
import numpy as np
import copy as copy
import matplotlib.pyplot as plt
import pandas as pd
import time 
starttime = time.time
import openrouteservice
from math import radians, cos, sin, asin, sqrt
import datetime
import math
from random import seed
from random import randint
seed(1)
from tqdm import tqdm
import random

#--------------------------------
#Загрузка параметров

average_car_speed = 110. #[km/h]
max_catchment_area_perimeter = 5. #[часы]
maximum_potential_access_egress_dist = average_car_speed * max_catchment_area_perimeter
max_accsess_egress_time = 2.

VoTime_access    = 67.5 #ценность времени до прибытия в ап(из дома в ап)
VoTime_waiting   = 75   #ценность времени ожидания
VoTime_invehicle = 50   #нормальное время в дороге
VoTime_transfer  = 50   #нормальное время во время трансфера
VoTime_egress    = 67.5 #ценность времени после выхода из ап(из ап до дома)

weight_acc = VoTime_access / VoTime_invehicle    #распределение весов и перевод их в нормальные единицы, так как ценность времени измеряется в Руб/ч
weight_wait = VoTime_waiting / VoTime_invehicle
weight_inv = 1.0
weight_tran = VoTime_transfer / VoTime_invehicle
weight_egr = VoTime_egress / VoTime_invehicle

time_access = 2. #time from leaving home to seat in plane
time_dec = 1.    #time from leaving plane to reach home
fac_dt = 1         #множитель времени, для поправки момента искревления пути
distance_acc = 50  #расстояние доступа
distance_dec = 50  #расстояние выхода
speed_kreys = 1800  #крейсерская скорость

#матрица времени для разных режимов[hours]
time_acc_m = [0.5, 0.25, 0.0]
time_wait_m = [2., 0.5, 0.0]
time_transfer_m = [2., 1.5, 0.0]
time_egres_m = [1., 0.5, 0.0]

time_wait = time_wait_m[0]

lowerboundary_distance = 300. #граничное условие для дистанции(мин дистанция для учета маршрута)
lowerboundary_duration = 3.   #граничное условие для времени(минимальная длительность)


fc_detour = [1., 1.09, 1.20] #plane, HSR, car коэффициент удлинения пути для различного транспорта
vehicle_speed = [1800., 220., 110.] #plane, HSR, car

daily_operational_hours = 22.0 #часы работы сети в день

#веса в функции полезности(определить xxx)
a1 = -1 * 2 #вес общего времени поездки (отрицательный – чем больше время, тем хуже). 
a2 = +1 * 3  #вес частоты отправлений (положительный – чем чаще рейсы, тем лучше).
a3 = -1 * -1 #вес пересечения границы (отрицательный – усложняет поездку).

'''--------------------------------------------------------
-----------------------   ШAГ1    -------------------------
---------    Загрузка данных o пассажиропотоках -----------'''

print('')
print('log. ШAГ 1  ------  Загрузка данных o пассажиропотоках')

# location = ([r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_vvo.xlsx', 
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_vko.xlsx',
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_svo.xlsx',
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_ovb.xlsx',
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_led.xlsx',
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_kzn.xlsx',
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_ikt.xlsx',
#              r'/Users/Dimka/Desktop/HSA network design/TNDP_Russia/avia_par_dme.xlsx]'])



location = ([r'TNDP_Russia/Source_data/avia_par_vvo.xlsx', 
             r'TNDP_Russia/Source_data/avia_par_vko.xlsx',
             r'TNDP_Russia/Source_data/avia_par_svo.xlsx',
             r'TNDP_Russia/Source_data/avia_par_ovb.xlsx',
             r'TNDP_Russia/Source_data/avia_par_led.xlsx',
             r'TNDP_Russia/Source_data/avia_par_kzn.xlsx',
             r'TNDP_Russia/Source_data/avia_par_ikt.xlsx',
             r'TNDP_Russia/Source_data/avia_par_dme.xlsx'])

sheet = (['avia_par_vvo',
          'avia_par_vko',
          'avia_par_svo', 
          'avia_par_ovb', 
          'avia_par_led', 
          'avia_par_kzn', 
          'avia_par_ikt', 
          'avia_par_dme'])

Airport_data = [['Unit', 'Msr', 'orig_ap', 'dest_ap', 'pax']]
Airport_data_frequency = [['Unit', 'Msr', 'orig_ap', 'dest_ap', 'flights']]

time.sleep(1)
print('')
print('log. Bытаскиваем данные от АП по паксам и полетам')

with tqdm(total=len(location), desc="Processing", bar_format="{l_bar}{bar} [ time left: {remaining} ]", position=0, leave=True) as pbar:  
    for u in range(len(location)):
        wb = load_workbook(location[u])
        ws = wb[sheet[u]]
        
        #загружаем файлы для обработки
        years = np.array([[i.value for i in j] for j in ws['A1':'LL1']]) 
        XX = np.zeros(len(years[0]), dtype=object)
        for i in range(len(XX)):
            XX[i] = str(copy.copy(years[0,i]))
        years = XX #years
        
        #ищем колонку sum
        column = np.where(years=='sum')[0][0]
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        
        #строим таблицу
        upperleft = str('A') + str('1')
        lowerright = str(alphabet[column]) + str(ws.max_row)
        Table = np.array([[i.value for i in j] for j in ws[upperleft:lowerright]])

        #убираем ненужные колонки + проверка
        correction = 0
        for column in range(len(Table[0])):
            if Table[0][column - correction] != 'xx':
                if Table[0][column - correction] != 'sum':
                    Table = np.delete(Table, column - correction, axis=1)
                    correction = correction + 1

        #собираем нужную информацию из табллиц
        for row in range(len(Table)):
            if Table[row][0] == 'PAS':
                if Table[row][1] == 'PAS_CRD':
                    Airport_data = np.append(Airport_data, [Table[row]], axis=0)
            if Table[row][0] == 'FLIGHT':
                if Table[row][1] == 'CAF_PAS':
                    Airport_data_frequency = np.append(Airport_data_frequency, [Table[row]], axis=0)
                    
        pbar.update(1)

time.sleep(1)
#save pax
print('')
print('log. Сохранение таблицы с паксами')
df = pd.DataFrame(Airport_data)
df.to_excel(excel_writer = "TNDP_Russia/Support_data/Airport_data.xlsx")

time.sleep(1)
#save freq
print('')
print('log. Сохранение таблицы с полетами')
df = pd.DataFrame(Airport_data_frequency)
df.to_excel(excel_writer = "TNDP_Russia/Support_data/Airport_data_frequency.xlsx")

print('')
print('Well done')


'''--------------------------------------------------------
-----------------------   ШAГ 2    -------------------------
------------    Очистка и модификация данных -----------'''

time.sleep(2)
print('')
print('log. ШAГ 2  ------  Очистка и модификация данных')
print('')
print('log. Загрузка данных по паксам')
wb_Airport_data = load_workbook(r'TNDP_Russia/Support_data/Airport_data.xlsx')
ws_Airport_data = wb_Airport_data['Sheet1']
Airport_data = np.array([[i.value for i in j] for j in ws_Airport_data['B2':'H14174']])  

time.sleep(1)
print('')
print('log. Загрузка данных по полетам')
wb_Airport_data_frequency = load_workbook(r'TNDP_Russia/Support_data/Airport_data_frequency.xlsx')
ws_Airport_data_frequency = wb_Airport_data_frequency['Sheet1']

data_rows = []
for row in ws_Airport_data_frequency.iter_rows(min_row=2, values_only=True):
    # Пропускаем пустые строки
    if any(cell is not None for cell in row):
        data_rows.append(list(row))

Airport_data_frequency = np.array(data_rows, dtype=object)

time.sleep(1)
print('')
print('log. Подгружаем данные об аэропортах и убираем лишние ячейки')

#Проверка
correction = 0
for row in range(len(Airport_data)):
    if Airport_data[row-correction][6] == ': ':
        Airport_data = np.delete(Airport_data, [row-correction], axis=0)
        correction = correction+1
        
correction = 0
for row in range(len(Airport_data)):
    if Airport_data[row-correction][3] == 'ZZZZ' or Airport_data[row-correction][5] == 'ZZZZ':
        Airport_data = np.delete(Airport_data, [row-correction], axis=0)
        correction = correction+1
        
time.sleep(1)
print('')
print( 'log. Идентификация оставшихся аэропортов')
#Создаем список уникальных ап
set_of_airports = set() 

for row in range(len(Airport_data)-1):
    # Добавляем только не-None значения
    if Airport_data[row+1][2] is not None and str(Airport_data[row+1][2]).strip() != '':
        set_of_airports.add(str(Airport_data[row+1][2]).strip())
    if Airport_data[row+1][3] is not None and str(Airport_data[row+1][3]).strip() != '':
        set_of_airports.add(str(Airport_data[row+1][3]).strip())

set_of_airports.discard('None')
set_of_airports.discard('')
# print(set_of_airports)
    
Num_airports  =  len(set_of_airports)

list_Airport_temp = sorted(list(set_of_airports))

list_of_airports = np.zeros(Num_airports, dtype=object)

for n in range(Num_airports):
    list_of_airports[n] = list_Airport_temp[n]


time.sleep(1)
print('')
print('log. Загружаем характеристики аэропортов.')
print('     по категории  : аэропорт по коду ИАТА') 
print('     топографически: страна')
print('     географически : широта, долгота')

wb_Aiport_info = load_workbook(r'TNDP_Russia/Source_data/airportinformation3.xlsx')
ws_Airport_info = wb_Aiport_info['airportdata']

name = np.array([[i.value for i in j] for j in ws_Airport_info['B2':'GPG2']]) 
XX = np.zeros(len(name[0]), dtype=object)
for i in range(len(XX)):
    XX[i] = str(copy.copy(name[0,i]))
name = XX #Название
# print("")
# print(name)

country = np.array([[i.value for i in j] for j in ws_Airport_info['B4':'GPG4']]) 
XX = np.zeros(len(country[0]), dtype=object)
for i in range(len(XX)):
    XX[i] = copy.copy(country[0,i])
country = XX #Страна
# print("")
# print(country)

IATA = np.array([[i.value for i in j] for j in ws_Airport_info['B7':'GPG7']]) 
XX = np.zeros(len(IATA[0]), dtype=object)
for i in range(len(XX)):
    XX[i] = str(copy.copy(IATA[0,i]))
IATA = XX #ИАТА-код
# print("")
# print(IATA)

Airport_longitude = np.array([[i.value for i in j] for j in ws_Airport_info['B8':'GPG8']]) 
XX = np.zeros(len(Airport_longitude[0]))
for i in range(len(XX)):
    XX[i] = str(copy.copy(Airport_longitude[0,i]))
Airport_longitude = XX #широта
# print("")
# print(Airport_longitude)

Airport_latitude = np.array([[i.value for i in j] for j in ws_Airport_info['B9':'GPG9']]) 
XX = np.zeros(len(Airport_latitude[0]))
for i in range(len(XX)):
    XX[i] = str(copy.copy(Airport_latitude[0,i]))
Airport_latitude = XX #долгота
# print("")
# print(Airport_latitude)

airport_information = np.full((Num_airports, 5),'n/a', dtype=object) #[IATA, name, coun, AP_lat, AP_lon]

#заполняем таблицу
for row in range(len(list_of_airports)):
    airport_information[row][0] = list_of_airports[row]
    for column in range(len(name)):
        if IATA[column] == airport_information[row][0]:
            # airport_information[row][1] = IATA[column]
            airport_information[row][1] = name[column]
            airport_information[row][2] = country[column]
            airport_information[row][3] = Airport_latitude[column]
            airport_information[row][4] = Airport_longitude[column]
            

#обновляем последовательность ап в коде IATA
for n in range(Num_airports):
    list_of_airports[n] = airport_information[n][0]

def Airport_index_num(a): #находим индекс аэропорта
    try:
        # Нормализуем входные данные (удаляем пробелы, приводим к верхнему регистру)
        search_code = str(a).strip().upper()
        
        # Ищем совпадение
        for idx, airport_code in enumerate(list_of_airports):
            if str(airport_code).strip().upper() == search_code:
                return idx
                
        # Если не найдено, выводим отладочную информацию
        return -1
        
    except Exception as e:
        print(f"Ошибка при поиске аэропорта '{a}': {str(e)}")
        return -1

time.sleep(1)
print('')
print ('log. Комбинируем характеристики аэропорта в одну матрицу')
# print ('')

#строим Origin-Destinaion матрицу
OD_matrix = np.zeros((len(list_of_airports), len(list_of_airports)),dtype=object)
OD_matrix_frequency = np.zeros((len(list_of_airports), len(list_of_airports)),dtype=object)

OD_matrix = np.zeros((len(list_of_airports), len(list_of_airports)), dtype=object)

for row in range(len(Airport_data)-1):
    orig = Airport_data[row+1][2]
    dest = Airport_data[row+1][3]
    # Проверяем, что orig и dest не None и не строка "None"
    if (orig is not None and dest is not None and 
        orig != "None" and dest != "None"):
        j = Airport_index_num(orig)
        i = Airport_index_num(dest)
        # Убедимся, что индексы валидны (не -1)
        if i != -1 and j != -1:
            pax = Airport_data[row+1][4]
            # Убедимся, что pax не None
            OD_matrix[i,j] = pax if pax is not None else 0



print('')
print('log. Матрица сделана')        
        
OD_matrix_float = np.zeros_like(OD_matrix, dtype=float)
for i in range(OD_matrix.shape[0]):
    for j in range(OD_matrix.shape[1]):
        try:
            OD_matrix_float[i,j] = float(OD_matrix[i,j])
        except (ValueError, TypeError):
            OD_matrix_float[i,j] = 0.0

#Заполняем OD матрицу частотой полетов
for row in range(len(Airport_data_frequency)-1):
    j = Airport_index_num(Airport_data_frequency[row+1][3])  # origin airport
    i = Airport_index_num(Airport_data_frequency[row+1][4])  # destination airport
    freq_y = Airport_data_frequency[row+1][5]  # frequency value

    if freq_y == ': ':
        freq_y = 0
    OD_matrix_frequency[i,j] = float(freq_y) / 365.  # convert yearly to daily
    if freq_y == 0 and OD_matrix[i,j] > 0:
        OD_matrix_frequency[i,j] = -1 * float('inf')  # маркер для отсутствующих данных
        

mirror_matrix = np.zeros((len(list_of_airports), len(list_of_airports)), dtype=float)
for i in range(len(mirror_matrix)):
    for j in range(len(mirror_matrix)):
        mirror_matrix[i,j] = np.maximum(OD_matrix_float[i,j], OD_matrix_float[j,i]) / 2.0 / 365.0
        mirror_matrix[j,i] = mirror_matrix[i,j]  # Не нужно copy.copy для numpy массивов           

mirror_matrix_freq = np.zeros((len(list_of_airports), len(list_of_airports)), dtype=object)
for i in range(len(mirror_matrix_freq)):
    for j in range(len(mirror_matrix_freq)):
        val_i = OD_matrix_frequency[i,j]
        val_j = OD_matrix_frequency[j,i]
        if val_i in [': ', 'n/a'] or val_j in [': ', 'n/a']:
            mirror_matrix_freq[i,j] = 'n/a'
        elif val_i == ':' or val_j == ':':
            mirror_matrix_freq[i,j] = ':'
        else:
            mirror_matrix_freq[i,j] = max(float(val_i), float(val_j)) / 2.0
        mirror_matrix_freq[j,i] = mirror_matrix_freq[i,j]
        
        
# print(OD_matrix_frequency)
        
OD_matrix_numeric = np.zeros_like(OD_matrix, dtype=float)

for i in range(OD_matrix.shape[0]):
    for j in range(OD_matrix.shape[1]):
        # Пропускаем специальные значения
        if OD_matrix[i,j] in [':', 'n/a', 'None', None]:
            OD_matrix_numeric[i,j] = 0.0
        else:
            try:
                OD_matrix_numeric[i,j] = float(OD_matrix[i,j])
            except (ValueError, TypeError):
                OD_matrix_numeric[i,j] = 0.0
                print(f"Нечисловое значение в [{i},{j}]: {OD_matrix[i,j]}")

            

sum_vertical = np.zeros((len(list_of_airports), 1))
for i in range(len(list_of_airports)):
    row_sum = 0.0
    for val in OD_matrix[i, :]:
        try:
            row_sum += float(val)
        except (ValueError, TypeError):
            pass  # Пропускаем нечисловые значения
    sum_vertical[i] = row_sum
    
   
sum_horizontal = np.zeros((len(list_of_airports), 1))
for j in range(len(list_of_airports)):
    row_sum = 0.0
    for val in OD_matrix[:, j]:
        try:
            row_sum += float(val)
        except (ValueError, TypeError):
            pass  # Пропускаем нечисловые значения
    sum_horizontal[j] = row_sum   
   

OD_matrix = OD_matrix.astype(object)
mirror_matrix = mirror_matrix.astype(object)

# Теперь можно присваивать строки
for i in range(len(OD_matrix)):
    for j in range(len(OD_matrix)):
        if OD_matrix[i,j] == 0:
            OD_matrix[i,j] = ':'
        if mirror_matrix[i,j] == 0:
            mirror_matrix[i,j] = ':'

# Сохраняем исходную версию массива
airport_information_OG = copy.copy(airport_information)

# Проверяем и корректируем размеры sum_horizontal
if sum_horizontal.shape[1] < airport_information.shape[1]:
    sum_horizontal = np.pad(sum_horizontal, ((0, 0), (0, airport_information.shape[1] - sum_horizontal.shape[1])), mode='constant')


sum_vertical = np.zeros((len(OD_matrix), 1), dtype=float)
for i in range(len(OD_matrix)):
    # Преобразуем строку OD_matrix[i, :] в массив float, заменяя нечисловые значения на 0
    row = np.array([float(x) if x not in [':', 'n/a', 'None'] else 0.0 for x in OD_matrix[i, :]])
    sum_vertical[i] = np.sum(row)

# Создаем sum_horizontal (суммы столбцов OD_matrix)
sum_horizontal = np.zeros((len(OD_matrix), 1), dtype=float)
for j in range(len(OD_matrix)):
    # Преобразуем столбец OD_matrix[:, j] в массив float, заменяя нечисловые значения на 0
    column = np.array([float(x) if x not in [':', 'n/a', 'None'] else 0.0 for x in OD_matrix[:, j]])
    sum_horizontal[j] = np.sum(column)

# Создаем основную часть таблицы (нормальную матрицу)
total_matrix = np.append(sum_vertical, OD_matrix, axis=1)
total_matrix = np.append(airport_information, total_matrix, axis=1)

# Транспонируем airport_information для верхней части
airport_information_transposed = np.transpose(airport_information)  # Форма: (5, Num_airports)

# Преобразуем sum_horizontal, чтобы его форма была (1, Num_airports)
sum_horizontal_reshaped = sum_horizontal.T  # Форма: (1, Num_airports)

# Объединяем airport_information_transposed и sum_horizontal_reshaped по оси 0
airport_information_transposed = np.append(airport_information_transposed, sum_horizontal_reshaped, axis=0)  # Форма: (6, Num_airports)

# Создаем левый верхний угол
left_hand_corner = np.full((airport_information_transposed.shape[0], airport_information_transposed.shape[0]), '', dtype=object)

# Объединяем left_hand_corner и airport_information_transposed по оси 1
upper_bar = np.append(left_hand_corner, airport_information_transposed, axis=1)


if upper_bar.shape[1] > total_matrix.shape[1]:
    # Дополняем total_matrix пустыми строками
    total_matrix = np.pad(
        total_matrix,
        ((0, 0), (0, upper_bar.shape[1] - total_matrix.shape[1])),
        mode='constant',
        constant_values=''
    )
else:
    # Дополняем upper_bar пустыми строками
    upper_bar = np.pad(
        upper_bar,
        ((0, 0), (0, total_matrix.shape[1] - upper_bar.shape[1])),
        mode='constant',
        constant_values=''
    )

# Объединяем верхнюю и нижнюю части
total_matrix = np.append(upper_bar, total_matrix, axis=0)

# Восстанавливаем исходный массив
airport_information = copy.copy(airport_information_OG)

time.sleep(1)
# Экспорт total_matrix в Excel
print('')
print('log. Сохранение полной матрицы')
df = pd.DataFrame(total_matrix).T
df.to_excel(excel_writer="TNDP_Russia/Support_data/DM_air_matrix.xlsx")




#-----------------------------------------------------------------------
#construct total output matrix which is mirrored
# Конструируем итоговую матрицу

sum_vertical_mirror = np.zeros((len(mirror_matrix), 1), dtype=float)
for i in range(len(mirror_matrix)):
    # Преобразуем строку mirror_matrix[i, :] в массив float, заменяя нечисловые значения на 0
    row = np.array([float(x) if x not in [':', 'n/a', 'None'] else 0.0 for x in mirror_matrix[i, :]])
    sum_vertical_mirror[i] = np.sum(row)

# Создаем sum_horizontal_mirror (суммы столбцов mirror_matrix)
sum_horizontal_mirror = np.zeros((len(mirror_matrix), 1), dtype=float)
for j in range(len(mirror_matrix)):
    # Преобразуем столбец mirror_matrix[:, j] в массив float, заменяя нечисловые значения на 0
    column = np.array([float(x) if x not in [':', 'n/a', 'None'] else 0.0 for x in mirror_matrix[:, j]])
    sum_horizontal_mirror[j] = np.sum(column)

# Создаем основную часть таблицы (зеркальную матрицу)
total_mirror_matrix = np.append(sum_vertical_mirror, mirror_matrix, axis=1)
total_mirror_matrix = np.append(airport_information, total_mirror_matrix, axis=1)

# Транспонируем airport_information для верхней части
airport_information_transposed = np.transpose(airport_information)  # Форма: (5, Num_airports)

# Преобразуем sum_horizontal_mirror, чтобы его форма была (1, Num_airports)
sum_horizontal_mirror_reshaped = sum_horizontal_mirror.T  # Форма: (1, Num_airports)

# Объединяем airport_information_transposed и sum_horizontal_mirror_reshaped по оси 0
airport_information_transposed = np.append(airport_information_transposed, sum_horizontal_mirror_reshaped, axis=0)  # Форма: (6, Num_airports)

# Создаем левый верхний угол
left_hand_corner_mirror = np.full((airport_information_transposed.shape[0], airport_information_transposed.shape[0]), '', dtype=object)

# Объединяем left_hand_corner_mirror и airport_information_transposed по оси 1
upper_bar_mirror = np.append(left_hand_corner_mirror, airport_information_transposed, axis=1)


if upper_bar_mirror.shape[1] > total_mirror_matrix.shape[1]:
    # Дополняем total_mirror_matrix нулями
    total_mirror_matrix = np.pad(
        total_mirror_matrix,
        ((0, 0), (0, upper_bar_mirror.shape[1] - total_mirror_matrix.shape[1])),
        mode='constant', constant_values=''
    )
else:
    # Дополняем upper_bar_mirror нулями
    upper_bar_mirror = np.pad(
        upper_bar_mirror,
        ((0, 0), (0, total_mirror_matrix.shape[1] - upper_bar_mirror.shape[1])),
        mode='constant', constant_values=''
    )

# Объединяем верхнюю и нижнюю части
total_mirror_matrix = np.append(upper_bar_mirror, total_mirror_matrix, axis=0)

# Восстанавливаем исходный массив
airport_information = copy.copy(airport_information_OG)

time.sleep(1)
# Экспорт total_mirror_matrix в Excel
print('')
print('log. Сохранение отзеркаленой матрицы')
df = pd.DataFrame(total_mirror_matrix).T
df.to_excel(excel_writer="TNDP_Russia/Support_data/DM_air_matrix_mirror.xlsx")            
  
time.sleep(1)          
#export total_matrix to excel
print('')
print('log. Сохранение полной матрицы полетов')

freq_matrix = np.zeros_like(OD_matrix_frequency, dtype=float)
for i in range(OD_matrix_frequency.shape[0]):
    for j in range(OD_matrix_frequency.shape[1]):
        value = OD_matrix_frequency[i, j]
        if value in [': ', 'n/a', 'None', None, ':']:
            freq_matrix[i, j] = 0.0  # или другое значение по умолчанию
        else:
            try:
                freq_matrix[i, j] = float(value)
            except (ValueError, TypeError):
                freq_matrix[i, j] = 0.0
                print(f"Нечисловое значение в [{i},{j}]: {value}")
                
                
df = pd.DataFrame(freq_matrix, index=list_of_airports, columns=list_of_airports)
df.to_excel(excel_writer="TNDP_Russia/Support_data/freq_air_matrix.xlsx", sheet_name="Flight Frequencies")


time.sleep(1)        
#export matrix to excel
print('')
print ('log. Сохранение зеркальной матрицы полетов')

mirror_freq_matrix = np.zeros_like(mirror_matrix_freq, dtype=float)
for i in range(mirror_matrix_freq.shape[0]):
    for j in range(mirror_matrix_freq.shape[1]):
        value = mirror_matrix_freq[i, j]
        if value in [':', 'n/a', 'None', None]:
            mirror_freq_matrix[i, j] = 0.0  # или другое значение по умолчанию
        else:
            try:
                mirror_freq_matrix[i, j] = float(value)
            except (ValueError, TypeError):
                mirror_freq_matrix[i, j] = 0.0
                print(f"Нечисловое значение в [{i},{j}]: {value}")
                
df_mirror = pd.DataFrame(mirror_freq_matrix, index=list_of_airports, columns=list_of_airports)
df_mirror.to_excel(excel_writer="TNDP_Russia/Support_data/freq_air_matrix_mirror.xlsx", sheet_name="Mirrored Flight Frequencies")

print('')
print ('Well done')



'''--------------------------------------------------------
-----------------------   ШAГ 3    -------------------------
------------    Очистка и модификация данных -----------'''

time.sleep(1)
print('')
print('log. ШAГ 3  ------  Расчёт расстояний между городами и аэропортами')

print('')
print('log. Добавляем данные по аэропортам, делая исключения:')
print('     добавляем только русские АП (Европы нет)')

print('')
print(airport_information)
print('')
# correction=0
# for x in range(len(airport_information)):   
#     if airport_information[x-correction][3] != 'Russia':
#         airport_information = np.delete(airport_information, x-correction, axis = 0)
#         mirror_matrix = np.delete(mirror_matrix, x-correction, axis = 0)
#         mirror_matrix = np.delete(mirror_matrix, x-correction, axis = 1)
#         mirror_matrix_freq = np.delete(mirror_matrix_freq, x-correction, axis = 0)
#         mirror_matrix_freq = np.delete(mirror_matrix_freq, x-correction, axis = 1)
#         correction = correction + 1
#import core cities

time.sleep(1)
print('')
print('log. Загружаем таблицу растояний ЖД путей российских городов ') 
print('')
wb = load_workbook(r'TNDP_Russia/Source_data/Core_cities_geography.xlsx')
ws_vertices = wb['Duration_road']

length_V = 8
range_len_V = range(8)

V = np.array([[i.value for i in j] for j in ws_vertices['G2':'N2']]) 
XX = np.zeros(length_V, dtype=object)
for i in range(len(XX)):
    XX[i] = V[0,i]
V = XX #vertex latitudes

V_country = np.array([[i.value for i in j] for j in ws_vertices['G3':'N3']]) 
XX = np.zeros(length_V, dtype=object)
for i in range(len(XX)):
    XX[i] = V_country[0,i]
V_country = XX #vertex latitudes

V_pop = np.array([[i.value for i in j] for j in ws_vertices['G6':'N6']]) 
XX = np.zeros(length_V)
for i in range(len(XX)):
    XX[i] = V_pop[0,i]
V_pop = XX #vertex populations

V_lat = np.array([[i.value for i in j] for j in ws_vertices['G4':'N4']]) 
XX = np.zeros(length_V)
for i in range(len(XX)):
    XX[i] = V_lat[0,i]
V_lat = XX #vertex latitudes

V_lon = np.array([[i.value for i in j] for j in ws_vertices['G5':'N5']]) 
XX = np.zeros(length_V)
for i in range(len(XX)):
    XX[i] = V_lon[0,i]
V_lon = XX #vertex longitudes 

print(airport_information)
#sort latitudes and longitudes based on airport information list
Airport_lat = np.zeros((len(airport_information)))
Airport_lon = np.zeros((len(airport_information)))
for row in range(len(Airport_lon)):
    Airport_lat[row] = airport_information[row][3]
    Airport_lon[row] = airport_information[row][4]
    
#build OD matrices for airport to city
City_to_Airport_Distance = np.zeros((len(V), len(airport_information))) #greater circle distance
City_to_Airport_Duration = np.zeros((len(V), len(airport_information))) #duration by car

#define haversine formula
def haversine(lat_i, lon_i, lat_j, lon_j):

    # convert decimal degrees to radians 
    lat_i, lon_i, lat_j, lon_j = map(radians, [lat_i, lon_i, lat_j, lon_j])

    # haversine formula 
    lat_delta = (lat_j - lat_i) 
    lon_delta = (lon_j - lon_i) 
    sr = sqrt(sin(lat_delta/2)**2 + cos(lat_i) * cos(lat_j) * sin(lon_delta/2)**2)
    R_earth = 6371.000 #radius earth [m]
    ds_gc = R_earth * 2 * asin(sr)
    return ds_gc #[km]


time.sleep(1)
#calculate greater circle distances matrix
print('')
print( 'log. Строим таблицу расстояний City - to - Airport' )
for i in range(len(City_to_Airport_Distance)):
    for j in range(len(City_to_Airport_Distance[0])):
        City_to_Airport_Distance[i,j] = haversine(V_lat[i],V_lon[i],airport_information[:,3][j],airport_information[:,4][j])

df = pd.DataFrame(City_to_Airport_Distance)
df.to_excel(excel_writer = "TNDP_Russia/Support_data/City_to_Airport_Distance.xlsx")



# '''--------------------------------------------------------
# -----------------------   ШAГ 4    -------------------------
# --- Расчет времени доступа между городами и аэропортами ---'''


time.sleep(1)
print('')
print('log. ШAГ 4  ------  Расчет времени доступа между городами и аэропортами')

print('')
print('log. Загрузка CIty - to - Airport расстояний через API(OpenRouteService)')

#define infeasible city-airport combinations
for i in range(len(City_to_Airport_Distance)):
    for j in range(len(City_to_Airport_Distance[0])):
        if City_to_Airport_Distance[i,j] > maximum_potential_access_egress_dist:
            City_to_Airport_Duration[i,j] = float('inf')

wb = load_workbook(r'TNDP_Russia/Support_data/City_to_Airport_Duration.xlsx')
ws = wb['Sheet1']
City_to_Airport_Duration = np.array([[i.value for i in j] for j in ws['B2':'I9']], dtype=object)

#determine access and egress times
saver = 0
counter = 0
for v in range(len(City_to_Airport_Duration)):
    for ap in range(len(City_to_Airport_Duration[0])):
        value = City_to_Airport_Duration[v, ap]
        if isinstance(value, str):
            if value.lower() == "inf":
                value = float('inf')
            else:
                value = float(value)
        else:
            value = float(value)
        if value < 0.000001:
            print('')
            print('от', ap, airport_information[ap, 1], '(аэропорта) до', v, V[v], '(города)')
            print(f"Coords for city={V[v]}, airport={airport_information[ap, 1]}: {((V_lon[v], V_lat[v]), (airport_information[ap, 3], airport_information[ap, 4]))}")
            distance = haversine(V_lat[v], V_lon[v], float(airport_information[ap, 3]), float(airport_information[ap, 4]))
            if distance * 1000 > 5900000:
                City_to_Airport_Duration[v, ap] = float('inf')
                print(f"Distance {distance:.2f} km exceeds limit, setting to inf")
                continue
            # if airport_information[ap, 1] == "Irkutsk Airport" and distance > 500:
            #     City_to_Airport_Duration[v, ap] = float('inf')
            #     print(f"Excluding Irkutsk Airport due to distance {distance:.2f} km")
            #     continue
            randomizer = 0
            if randomizer == 0:
                coords = ((V_lon[v], V_lat[v]), (airport_information[ap, 4], airport_information[ap, 3]))
                client = openrouteservice.Client(key='5b3ce3597851110001cf62481a403ebd11a94d8fbf2320c2f8eac293')
                try:
                    routes = client.directions(
                        coordinates=coords,
                        profile='driving-car',
                        radiuses=[30000, 30000],  # Increased to 30 km
                        format='geojson',
                        validate=False,
                    )
                    duration = routes['features'][0]['properties']['segments'][0]['duration'] / 60 / 60
                    if duration < max_catchment_area_perimeter:
                        City_to_Airport_Duration[v, ap] = duration
                    else:
                        City_to_Airport_Duration[v, ap] = float('inf')
                except openrouteservice.exceptions.ApiError as e:
                    print(f"API Error for city={V[v]}, airport={airport_information[ap, 1]}: {e}. Setting to inf.")
                    City_to_Airport_Duration[v, ap] = float('inf')
                except Exception as e:
                    print(f"Unexpected error for city={V[v]}, airport={airport_information[ap, 1]}: {e}. Setting to inf.")
                    City_to_Airport_Duration[v, ap] = float('inf')
            counter = counter + 1
            saver = saver + 1
            print(counter, '(', v, ',', ap, ')')
            if saver == 19:
                df_dur = pd.DataFrame(City_to_Airport_Duration)
                df_dur.to_excel(excel_writer="TNDP_Russia/Suppot_data/City_to_Airport_Duration.xlsx")

df_dur = pd.DataFrame(City_to_Airport_Duration)
df_dur.to_excel(excel_writer="TNDP_Russia/Support_data/City_to_Airport_Duration.xlsx")

df_dur = pd.DataFrame(airport_information)
df_dur.to_excel(excel_writer="TNDP_Russia/Support_data/airport_information.xlsx")

time.sleep(1)
#""" #only activated in first run
print('')
print('log. Загрузка City - to - Airport матрицы расстояний')
wb = load_workbook(r'TNDP_Russia/Support_data/City_to_Airport_Duration.xlsx')
ws  =    wb['Sheet1']
City_to_Airport_Duration = np.array([[i.value for i in j] for j in ws['B2':'I9']],dtype=object)


'''--------------------------------------------------------
-----------------------   ШAГ 5    -------------------------
----------- Разработка всех возможных маршрутов ------------'''



time.sleep(1)
print('')
print('log. ШAГ 5  ------  Разработка всех возможных маршрутов')

print('')
print('log. Загружаем данные о расстояниях между городами (центры)' )
wb_road_distances = load_workbook(r'TNDP_Russia/Source_data/Core_cities_geography.xlsx')
ws_road_distances = wb_road_distances['Distance_road']
# road_distance = np.array([[i.value for i in j] for j in ws_road_distances['G7':'N14']], dtype=float)



#-------------------------------------#
road_distance = []
for row in ws_road_distances['G7':'N14']:
    row_data = []
    for cell in row:
        value = cell.value
        if value is not None:
            # Replace non-breaking space (\xa0) and other whitespace with a regular space, then convert
            value = str(value).replace(u'\xa0', ' ').replace(' ', '').strip()
            try:
                value = float(value) if value else 0.0
                row_data.append(value)
            except ValueError:
                row_data.append(0.0)  # Default to 0 for unconvertible values
                print(f"Non-numeric value at {cell.coordinate}: {cell.value}, replaced with 0.0")
        else:
            row_data.append(0.0)
    road_distance.append(row_data)

road_distance = np.array(road_distance, dtype=float)
#-------------------------------------#


time.sleep(1)
print('')
print('log. Загружаем данные о временных интервалах между городами' )
wb_road_duration = load_workbook(r'TNDP_Russia/Source_data/Core_cities_geography.xlsx')
ws_road_duration = wb_road_distances['Duration_road']
# road_duration = np.array([[i.value for i in j] for j in ws_road_duration['G7':'N14']], dtype=float)

#-------------------------------------#
wb_road_duration = load_workbook(r'TNDP_Russia/Source_data/Core_cities_geography.xlsx')
ws_road_duration = wb_road_duration['Duration_road']  # Fixed typo: use wb_road_duration
road_duration = []
for row in ws_road_duration['G7':'N14']:
    row_data = []
    for cell in row:
        value = cell.value
        if value is not None:
            value = str(value).replace(u'\xa0', ' ').replace(' ', '').strip()
            try:
                value = float(value) if value else 0.0
                row_data.append(value)
            except ValueError:
                row_data.append(0.0)
                print(f"Non-numeric value at {cell.coordinate}: {cell.value}, replaced with 0.0")
        else:
            row_data.append(0.0)
    road_duration.append(row_data)

road_duration = np.array(road_duration, dtype=float)
#-------------------------------------#

# print('')
# print("log. Road distances (km):")
# print(road_distance)
# print('')
# print("log. Road durations (hours):")
# print(road_duration)


time.sleep(1)
print('')
print( 'log. Подсчитываем полетное время между аэропортами')
def func_t_inv_air(ds_gc_ij):
    
    #рассчет полетного времени
    t_inv_air = time_access + (ds_gc_ij * fac_dt - distance_acc - distance_dec) / (speed_kreys) + time_dec 
        
    return t_inv_air 

T_inv_air = np.full((len(mirror_matrix), len(mirror_matrix)), float('inf'))
for i in range(len(airport_information)):
    for j in range(len(airport_information)):
        if mirror_matrix[i,j] != ':':
            AP_to_AP_dis_gc = haversine(float(airport_information[i, 3]), float(airport_information[i, 4]),
                                       float(airport_information[j, 3]), float(airport_information[j, 4]))
            T_inv_air[i,j] = func_t_inv_air(AP_to_AP_dis_gc)
            
#verification of flight duration  
#print airport_information[192][0], airport_information[367][0], str(datetime.timedelta(seconds=T_inv_air[192,367]*60*60))
#print airport_information[192][0], airport_information[286][0], str(datetime.timedelta(seconds=T_inv_air[192,286]*60*60))
#print airport_information[192][0], airport_information[265][0], str(datetime.timedelta(seconds=T_inv_air[192,265]*60*60))
#print airport_information[192][0], airport_information[281][0], str(datetime.timedelta(seconds=T_inv_air[192,281]*60*60))


time.sleep(1)
#construction flight possibilities for each city-to-city OD pair
print('')
print( 'log. Определение городов в зоне досягаемости аэропортов (catchment area)' )                     
#build scatter lists for each city V
scatter_list_City_to_Airport = np.zeros((length_V,1), dtype=object)
for v in range_len_V:
    scatter_AP = list()
    for ap in range(len(airport_information)):
        if City_to_Airport_Duration[v][ap] != 'inf' and City_to_Airport_Duration[v][ap] < max_accsess_egress_time :
            scatter_AP = np.append(scatter_AP, ap, axis=None)
    #print v, V[v], scatter_AP
    scatter_list_City_to_Airport[v][0] = scatter_AP

time.sleep(1)
#build flight-options
print('')
print('log. Определяем возможные траектории полета для каждой пары городов')   
flight_options = np.zeros((length_V, length_V), dtype=object)
for i in range_len_V:
    for j in range_len_V:
        if i != j:
            flight_options[i,j] = np.zeros((len(scatter_list_City_to_Airport[i][0]),len(scatter_list_City_to_Airport[j][0])))

for i in range_len_V:
    for j in range_len_V: 
        if i != j:
            for x in range(len(flight_options[i,j])):
                for y in range(len(flight_options[i,j][0])):
                    ap1 = int(scatter_list_City_to_Airport[i][0][x])
                    ap2 = int(scatter_list_City_to_Airport[j][0][y])
                    flight_options[i,j][x,y] = T_inv_air[ap1,ap2]
                    


""""---------------------------------------------------------------------------
--------------------------------     ШАГ 6.    -------------------------------
------------------ Cравнение с другими способами передвижения-----------------
----------------------------------------------------------------------------"""


time.sleep(1)
print('')
print('log. ШAГ 6  ------  Cравнение с другими способами передвижения')

#Рассчитать общее взвешенное время полета на самолете.
def func_t_trip_air_weighted(time_access,time_wait,t_in_vehicle,t_egress):

    t_trip_air = (time_access * weight_acc) + (time_wait * weight_wait) + (t_in_vehicle * weight_inv) + (t_egress*weight_egr) #no transfers needed

    return t_trip_air #[h]


#рассчитать общее взвешенное время поездки на самолете, но с явной привязкой к паре городов по i/j
def func_t_trip_air_estimated(i,j,time_access,time_wait,t_in_vehicle,t_egress):

    t_trip_air_estimated = (time_access * weight_acc) + (time_wait * weight_wait) + (t_in_vehicle * weight_inv) + (t_egress*weight_egr) #no transfers needed

    return t_trip_air_estimated #[h]


#Рассчитать общее взвешенное время поездки на автомобиле между городами
def func_t_trip_car_estimated(i,j):

    t_trip_car_estimated = (road_duration[i,j]*weight_inv)

    return t_trip_car_estimated #[h]


#Рассчитать вероятность выбора авиатранспорта по сравнению с автомобилем с использованием модели случайного сожаления (Random Regret Model, RRM).
def func_modalsplit_estimate_air(tt_air,tt_car): #find mode specific regret values
    
    #Regret value determination
    R_air  = np.log(1.0 + math.e**(-0.01 * 60 * (tt_car - tt_air))) - np.log(2)
    R_car  = np.log(1.0 + math.e**(-0.01 * 60 * (tt_air - tt_car))) - np.log(2)

    #Mode probability determination
    P_air  = math.e**(-R_air) / (math.e**(-R_air) + math.e**(-R_car))
    
    return P_air


#Рассчитать "сожаление" (regret) для конкретного маршрута полета по сравнению с другими маршрутами для той же пары городов.
def func_regret(trip_time,trip_surrounding): #find flight specific regret values
    
    dif_sum = 0
    for tt_x in range(len(trip_surrounding)):
        dif = math.e**(-0.01 * 60 * (float(trip_surrounding[tt_x])- float(trip_time)))
        dif_sum = dif_sum + dif
    R = np.log( 1.0 + dif_sum ) - np.log(2)
    return R


#Рассчитать вероятность выбора конкретного маршрута полета на основе значений сожаления.
def func_probability(R_x,R_sum_x): #find flight specific probabilities
    
    P = math.e**(-1 * float(R_x)) / float(R_sum_x)
    
    return P 

#Рассчитать полезность (utility) конкретного маршрута полета с учетом времени доступа, выхода, расстояния и "border crossing" 
#(пересечения границ, в данном случае не используется, так как все города в России).
def func_utility_flight(tt_a, tt_e, DST, BC_ap1, BC_ap2): #find flight specific probabilities
    
    # U = a1*( ( tt_a*(a2*BC_ap1))*(tt_e*(a2*BC_ap2) ) ) + (a3*DST)
    U = a1*( ( tt_a*(a2))*(tt_e*(a2) ) ) + (a3*DST)
    return U 


#Рассчитать вероятность выбора маршрута на основе полезности (utility maximization).
def func_utility_maximisation_probability(V_x,V_sum_x): #find flight specific probabilities
    
    P = math.e**(float(V_x)) / float(V_sum_x)
    
    return P 


# print('')
# print("log. Mirror matrix sample (first 5x5):")
# print(mirror_matrix[:8, :8])
# print('')
# print("log. T_inv_air sample (first 5x5):")
# print(T_inv_air[:8, :8])
# print('')

time.sleep(1)
print('')
time.sleep(1)
print("log. Список достижений от аэропортов до городов :")
print('')
for v in range_len_V:
    print(f"City {V[v]}: {scatter_list_City_to_Airport[v][0]}")




""""---------------------------------------------------------------------------
--------------------------------     STEP 7.    -------------------------------
--------------------  Construction of connectivity list -----------------------
-----------------------------------------------------------------------------"""

time.sleep(1)
print('')
print('log. ШAГ 7  ------  Строим таблицу взаимосвязей')


#""" Only actived when adjusting certain parameters

#make average flight duration matrix
Matrix_avg_flight = np.full((length_V, length_V), float('inf')) 

time.sleep(1)
#make connectivity list
print('')
print('log. Строим таблицу взаимосвязей, включая:')
print('  ','- город вылета, аэропорт вылета, аэропорт назначения, город назначения')
print('  ','-','время до прибытия в ап, время от аэропорта до места назначения, время полета')
print('  ','-','взвешенные времена, относящиея к ValueOfTime')
print('  ','-','относительная величина сожаления, вероятность выбора пути и потенциал')
print('')
connectivity_list = [['i','access','ap1','flight','ap2','egress','j','dur_tot','dur_tot_wei','R','P','potential','pax','pax_ij_tot','marketshare_ij','freq','cb_ap1','cb_ap2','Utility', 'prob. MNL']]
with tqdm(total=length_V, desc="Processing", bar_format="{l_bar}{bar} [ time left: {remaining} ]", position=0, leave=True) as pbar:
    for i in range_len_V:
        pbar.update(1)
        for j in range_len_V: 
            if i != j:
                #if (i == 109 and j == 122) or (i == 112 and j == 122) :
                if 1 == 1:
                    if road_distance[i,j] > lowerboundary_distance or road_duration[i,j] > lowerboundary_duration:
                        #make list of feasible routes
                        feasible_routes = list()
                        for x in range(len(flight_options[i,j])):
                            for y in range(len(flight_options[i,j][0])):
                                #fill in already available data concerning feasible trip characteristics
                                if flight_options[i,j][x,y] != float('inf'):
                                    trip = [i,'t_acc',int(scatter_list_City_to_Airport[i][0][x]),flight_options[i,j][x,y],int(scatter_list_City_to_Airport[j][0][y]),'t_egr',j,'t_tot', 't_weight','R', 'P', 'pot', 'pax','pax_ij_tot','marketshare_ij','freq',0.0,0.0,'Utility', 'prob. MNL']
                                    trip[1] = City_to_Airport_Duration[i,trip[2]]
                                    trip[5] = City_to_Airport_Duration[j,trip[4]]
                                    trip[7] = trip[1] + trip[3] + trip[5] + time_wait
                                    trip[8] = func_t_trip_air_weighted(trip[1],time_wait,trip[3],trip[5])
                                    trip[15] = mirror_matrix_freq[trip[2],trip[4]]
                                    if V_country[int(trip[0])] != airport_information[trip[2]][4]:
                                        trip[16] = 1.0 
                                    if airport_information[trip[4]][4] != V_country[int(trip[6])]:
                                        trip[17] = 1.0
                                    #append this trip to the list of feasbile routes
                                    if len(feasible_routes) == 0:
                                        feasible_routes = [trip]
                                    else:
                                        feasible_routes = np.append(feasible_routes, [trip], axis=0)
                        #calculate regret value
                        if len(feasible_routes) > 1:
                            #print('')
                            for trip in range(len(feasible_routes)):
                                trip_time = feasible_routes[trip,8]   
                                trip_surrounding = np.delete(feasible_routes[:,8], trip)
                                R = func_regret(trip_time,trip_surrounding)
                                feasible_routes[trip,9] = R
                                tt_a = float(feasible_routes[trip,1])
                                tt_e = float(feasible_routes[trip,5])
                                if feasible_routes[trip,15] == 'n/a': #estimate value when frequency is missing
                                    feasible_routes[trip,15] = mirror_matrix[int(feasible_routes[trip,2]),int(feasible_routes[trip,4])] / 365 / 150 #150 estimate average flight
                                if feasible_routes[trip,15] == ':': #estimate value when frequency is missing
                                    feasible_routes[trip,15] = mirror_matrix[int(feasible_routes[trip,2]),int(feasible_routes[trip,4])] / 365 / 150 #150 estimate average flight
                                DST = (daily_operational_hours / float(feasible_routes[trip,15])) / 4.0
                                BC_ap1 = float(feasible_routes[trip,16]) 
                                BC_ap2 = float(feasible_routes[trip,17])
                                U = func_utility_flight(tt_a, tt_e, DST, BC_ap1, BC_ap2)
                                feasible_routes[trip,18] = U
                            #summate R value as preparation for P calculations
                            R_sum_x = 0
                            for trip in range(len(feasible_routes)):
                                R_sum_x = R_sum_x + math.e**(-1 * float(feasible_routes[trip,9]))
                                
                            U_sum_x = 0
                            for trip in range(len(feasible_routes)):
                                U_sum_x = U_sum_x + math.e**(float(feasible_routes[trip,18]))
                                #print '(',round(float(feasible_routes[trip,8]),1), round(float(feasible_routes[trip,15]),1), int(float(feasible_routes[trip,16]) + float(feasible_routes[trip,17])),')' ,
                            #print('')
                            #calculate probability value
                            for trip in range(len(feasible_routes)):
                                p = func_probability(feasible_routes[trip,9],R_sum_x)
                                feasible_routes[trip,10] = p
                                #print round(p,3),
                            #print('')
                            for trip in range(len(feasible_routes)):
                                p_mnl = func_utility_maximisation_probability(feasible_routes[trip,18],U_sum_x)
                                feasible_routes[trip,19] = p_mnl
                                #print round(p_mnl,3),
                            #print('')
                            #calculate average flight trip
                            flight_avg = [0,time_wait_m[0],0,0] #'t_access',t_wait,'t_inv','t_egress'
                            for trip in range(len(feasible_routes)):
                                flight_avg[0] = flight_avg[0] + (float(feasible_routes[trip,1]) * float(feasible_routes[trip,19])) #access time * probability
                                flight_avg[2] = flight_avg[2] + (float(feasible_routes[trip,3]) * float(feasible_routes[trip,19])) #inv. time * probability
                                flight_avg[3] = flight_avg[3] + (float(feasible_routes[trip,5]) * float(feasible_routes[trip,19])) #egress * probability
                            #estimate travel times using other modes
                            tt_air  = func_t_trip_air_estimated(i, j, flight_avg[0], flight_avg[1], flight_avg[2], flight_avg[3] )
                            tt_car  = func_t_trip_car_estimated(i, j                                                             )
                            #save average flight time for later works
                            Matrix_avg_flight[i,j] = tt_air
                            #estimate modal split for air travel when comparing to other modes
                            MS_air_estimate = func_modalsplit_estimate_air(tt_air,tt_car)
                            #translate probability into potential
                            for trip in range(len(feasible_routes)):
                                potential = MS_air_estimate * ( V_pop[i] * V_pop[j] ) / (road_distance[i,j] / fc_detour[2]) * float(feasible_routes[trip,10])
                                feasible_routes[trip,11] = potential
                            #append trip data to connectivity list
                            for trip in range(len(feasible_routes)):
                                connectivity_list = np.append(connectivity_list, [feasible_routes[trip]], axis=0)
                        if len(feasible_routes) == 1:
                            #give values to R and P
                            feasible_routes[0][9] = 1.00
                            feasible_routes[0][10] = 1.00
                            #define only flight as average flight
                            flight_avg = [float(feasible_routes[0][1]),time_wait_m[0],float(feasible_routes[0][3]),float(feasible_routes[0][5])]
                            #define competing travel times 
                            tt_air  = func_t_trip_air_estimated( i, j, flight_avg[0], flight_avg[1], flight_avg[2], flight_avg[3] )
                            tt_car  = func_t_trip_car_estimated( i, j                                                             )
                            #save average flight time for later works
                            Matrix_avg_flight[i,j] = tt_air
                            #estimate modal split for air travel when comparing to other modes
                            MS_air_estimate = func_modalsplit_estimate_air(tt_air,tt_car)
                            #translate probability into potential
                            feasible_routes[0][11] = MS_air_estimate * ( V_pop[i] * V_pop[j] ) / (road_distance[i,j] / fc_detour[2]) 
                            #append trip data to connectivity list
                            connectivity_list = np.append(connectivity_list, [feasible_routes[0]], axis=0)
                            #print MS_air_estimate
            
#delete header row from connectivity list in order to calculate on it
connectivity_list = np.delete(connectivity_list, 0, axis = 0)

time.sleep(1)
print('')
print(Matrix_avg_flight)

""""---------------------------------------------------------------------------
--------------------------------     STEP 8.    -------------------------------
-------------  Перевод спроса на авиаперевозки с аэропортов на города ---------
-----------------------------------------------------------------------------"""

time.sleep(1)
print('')
print('log. ШAГ 8  ------  Перевод спроса на авиаперевозки с аэропортов на города')

#make airport numbers integers again
DEMAND_AIR = np.zeros((length_V, length_V))
Matrix_avg_flight_2 = np.zeros((length_V, length_V), dtype=object)
for i in range_len_V:
    for j in range_len_V:
        Matrix_avg_flight_2[i,j] = [0,0,0,0,0,0] #'t_access',t_wait,'t_inv','t_egress', t_tot, t_w_tot

time.sleep(1)
APx_to_APy_list_total = [['i','ac','ap1','fl','ap2','eg','j','dt','dtw','R','P','pot','pax','pax_ij_tot','marketshare_ij','freq','cb_ap1','cb_ap2','Utility', 'prob. MNL']]
print('')
print( 'log. Переводим спрос на авиаперевозки с аэропортов на города:')
with tqdm(total=len(airport_information), desc="log. Processing", bar_format="{l_bar}{bar} [ time left: {remaining} ]",position=0, leave=True) as pbar:
    for x in range(len(airport_information)):
        pbar.update(1)
        #for z in tqdm(range(len(airport_information))): #loading bar
        for y in range(len(airport_information)):
            if mirror_matrix[x,y] != ':':
                #lookup for all possible flights
                APx_to_APy_list = [['i','ac','ap1','fl','ap2','eg','j','dt','dtw','R','P','pot','pax','pax_ij_tot','marketshare_ij','freq','cb_ap1','cb_ap2','Utility', 'prob. MNL']]
                for row in range(len(connectivity_list)):
                    if int(float(connectivity_list[row][2])) == int(x) and int(float(connectivity_list[row][4])) == int(y):
                            APx_to_APy_list = np.append(APx_to_APy_list, [connectivity_list[row]], axis=0)
                APx_to_APy_list = np.delete(APx_to_APy_list, 0, axis=0)
                if len(APx_to_APy_list) != 0:
                    sum_potential = 0
                    for trip in range(len(APx_to_APy_list)):
                        sum_potential = sum_potential + float(APx_to_APy_list[trip,11])
                    #print APx_to_APy_list, sum_potential
                    for trip in range(len(APx_to_APy_list)):
                        marketshare = float(APx_to_APy_list[trip,11]) / sum_potential
                        i = int(float(APx_to_APy_list[trip,0])); j = int(float(APx_to_APy_list[trip,6]))
                        #print marketshare, i, j
                        DEMAND_AIR[i,j] = DEMAND_AIR[i,j] + (float(marketshare) * float(mirror_matrix[x,y]))
                        APx_to_APy_list[trip,12] = (float(marketshare) * float(mirror_matrix[x,y]))
                        #Matrix_avg_flight_2[i,j][0] = Matrix_avg_flight_2[i,j][0] + (float(marketshare) * float(APx_to_APy_list[trip,1]))
                        #Matrix_avg_flight_2[i,j][1] = Matrix_avg_flight_2[i,j][1] + (float(marketshare) * t_wait)
                        #Matrix_avg_flight_2[i,j][2] = Matrix_avg_flight_2[i,j][2] + (float(marketshare) * float(APx_to_APy_list[trip,3]))
                        #Matrix_avg_flight_2[i,j][3] = Matrix_avg_flight_2[i,j][3] + (float(marketshare) * float(APx_to_APy_list[trip,5]))
                APx_to_APy_list_total = np.append(APx_to_APy_list_total, APx_to_APy_list, axis=0)
APx_to_APy_list_total = np.delete(APx_to_APy_list_total, 0, axis=0)                

time.sleep(1)
print('')
print( round(DEMAND_AIR[1,0],0), V[1], V[0])
print (round(DEMAND_AIR[2,0],0), V[2], V[0])
print( round(DEMAND_AIR[3,0],0), V[3], V[0])
print( round(DEMAND_AIR[4,0],0), V[4], V[0])
print( round(DEMAND_AIR[5,0],0), V[5], V[0])

for row in range(len(APx_to_APy_list_total)):
    APx_to_APy_list_total[row][13] = DEMAND_AIR[int(APx_to_APy_list_total[row][0]), int(APx_to_APy_list_total[row][6])]
    APx_to_APy_list_total[row][14] = float(APx_to_APy_list_total[row][12]) / float(APx_to_APy_list_total[row][13])

for row in range(len(APx_to_APy_list_total)):
    i = int(APx_to_APy_list_total[row][0])
    j = int(APx_to_APy_list_total[row][6])
    marketshare_ixyj = float(APx_to_APy_list_total[row][14])
    Matrix_avg_flight_2[i,j][0] = Matrix_avg_flight_2[i,j][0] + (marketshare_ixyj * float(APx_to_APy_list_total[row,1]))
    Matrix_avg_flight_2[i,j][1] = Matrix_avg_flight_2[i,j][1] + (marketshare_ixyj * time_wait)
    Matrix_avg_flight_2[i,j][2] = Matrix_avg_flight_2[i,j][2] + (marketshare_ixyj * float(APx_to_APy_list_total[row,3]))
    Matrix_avg_flight_2[i,j][3] = Matrix_avg_flight_2[i,j][3] + (marketshare_ixyj * float(APx_to_APy_list_total[row,5]))

for i in range_len_V:
    for j in range_len_V:
        Matrix_avg_flight_2[i,j][4] = Matrix_avg_flight_2[i,j][0] + Matrix_avg_flight_2[i,j][1] + Matrix_avg_flight_2[i,j][2] + Matrix_avg_flight_2[i,j][3]
        Matrix_avg_flight_2[i,j][5] = func_t_trip_air_weighted(Matrix_avg_flight_2[i,j][0], Matrix_avg_flight_2[i,j][1], Matrix_avg_flight_2[i,j][2], Matrix_avg_flight_2[i,j][3])

for i in range_len_V:
    for j in range_len_V:
        if np.sum(Matrix_avg_flight_2[i,j]) < 0.00001:
            Matrix_avg_flight_2[i,j] = [float('inf'),float('inf'),float('inf'),float('inf'),float('inf'),float('inf')]
        
print('') 
print( Matrix_avg_flight_2[0,0] ) 
print( Matrix_avg_flight_2[1,0] ) 
print( Matrix_avg_flight_2[2,0])   
print( Matrix_avg_flight_2[3,0]  )
print( Matrix_avg_flight_2[4,0])  
print( Matrix_avg_flight_2[5,0])  
print( Matrix_avg_flight_2[6,0])
print( Matrix_avg_flight_2[7,0])       

time.sleep(1)
print('')
print('log. Сохраняем матрицу переведенного спроса с аэропортов на города')
#export total_matrix to excel
df = pd.DataFrame(DEMAND_AIR)
df.to_excel(excel_writer = "TNDP_Russia/DEMAND_AIR.xlsx")     



""""---------------------------------------------------------------------------
--------------------------------     STEP 9.    -------------------------------
------------------------  Перекладываем на общий спрос -------------------------
-----------------------------------------------------------------------------"""
 
time.sleep(1)
print('')
print('log. ШAГ 9  ------  Перекладываем на общий спрос')

time.sleep(1)
print('')
print('log. Трансляция спроса авиаперевозок на общий спрос' )
           
def MS_air_predictor(distance):
    if distance <= 200.:
        MS_air = float('inf')
    if distance > 200. and distance < 1500.:
        MS_air =((1.4940581931E-12*distance**4) - (4.7257341849021E-09*distance**3) + (4.59662481788692E-6*distance**2) - (5.0276000274593E-4*distance))
        MS_air = min(MS_air, 1.0)  # Ограничиваем долю значением 1
    if distance >= 1500.:
        MS_air = 1.00
    
    return MS_air

DEMAND_TOTAL = np.zeros((length_V, length_V))
for i in range_len_V:
    for j in range_len_V:
        if i != j:
            DEMAND_TOTAL[i,j] = DEMAND_AIR[i,j] / MS_air_predictor(road_distance[i,j])
            #tt_air  = Matrix_avg_flight_2[i,j][5]
            #tt_car  = func_t_trip_car_estimated( i, j    )
            #if i == 109 and j == 122:
                #print 'v_i v_j   |   MSair MSair   |   air car PT'
                #print i,j, '  |  ' , round(MS_air_predictor(road_distance[i,j]),3), round(MS_air_estimate,3), '  |  ' , round(tt_air,2), round(tt_car,2)
            #if i == 110 and j == 122:
                #print i,j, '  |  ' , round(MS_air_predictor(road_distance[i,j]),3), round(MS_air_estimate,3), '  |  ' , round(tt_air,2), round(tt_car,2)
            #if i == 111 and j == 122:
                #print i,j, '  |  ' , round(MS_air_predictor(road_distance[i,j]),3), round(MS_air_estimate,3), '  |  ' , round(tt_air,2), round(tt_car,2)
            #if i == 112 and j == 122:
                #print i,j, '  |  ' , round(MS_air_predictor(road_distance[i,j]),3), round(MS_air_estimate,3), '  |  ' , round(tt_air,2), round(tt_car,2)
            #if i == 113 and j == 122:
                #print i,j, '  |  ' , round(MS_air_predictor(road_distance[i,j]),3), round(MS_air_estimate,3), '  |  ' , round(tt_air,2), round(tt_car,2)            
            #print round(DEMAND_TOTAL[i,j]), round(DEMAND_AIR[i,j]), MS_air_predictor(road_distance[i,j])
#'''

time.sleep(1)
print("")
print(DEMAND_TOTAL)
#export total_matrix to excel
df = pd.DataFrame(DEMAND_TOTAL)
df.to_excel(excel_writer = "TNDP_Russia/DEMAND_TOTAL.xlsx")     
          

# print('')
# print( round(DEMAND_TOTAL[109,122]), V[109], V[122])
# print( round(DEMAND_TOTAL[110,122]), V[110], V[122])
# print( round(DEMAND_TOTAL[111,122]), V[111], V[122])
# print( round(DEMAND_TOTAL[112,122]), V[112], V[122])
# print( round(DEMAND_TOTAL[113,122]), V[113], V[122] )
 
#'''